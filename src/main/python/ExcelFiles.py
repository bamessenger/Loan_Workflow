import pandas as pd
import numpy as np
import pathlib as p
import win32com.client as win32

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


class XLFile:

    def fileRead(self, encompPath):
        fileExt = p.Path(encompPath).suffix
        if fileExt == '.csv':
            self.encmpDataAll = pd.read_csv(encompPath, header=0)
        else:
            self.encmpDataAll = pd.read_excel(encompPath, engine='openpyxl')
            self.encmpDataAll.columns = self.encmpDataAll.columns.str.replace(
                ' ', '').str.replace('MilestoneDate-', '')
            self.encmpDataAll['LoanStatus'] = np.where(
                self.encmpDataAll.ClosedDate.isnull(), 'Open', 'Closed')
            self.encmpDataAllAct = self.encmpDataAll.assign(DateType='Actual')
            self.encmpDataAllAct[['ApplicationDate', 'Disclosures',
                                  'Processing', 'submittal', 'Approval',
                                  'ConditionSubmission', 'ClearToClose',
                                  'OS-FinancingContingency',
                                  'LockExpirationDate', 'EstClosingDate',
                                  'ClosedDate']] = \
                self.encmpDataAllAct[['ApplicationDate', 'Disclosures',
                                  'Processing', 'submittal', 'Approval',
                                  'ConditionSubmission', 'ClearToClose',
                                  'OS-FinancingContingency',
                                  'LockExpirationDate', 'EstClosingDate',
                                  'ClosedDate']].apply(pd.to_datetime)
            self.encmpDataOpen = self.encmpDataAll[
                self.encmpDataAll.ClosedDate.isnull()]
            self.encmpDataOpen = self.encmpDataOpen.reset_index()

    def excelWrite(self, wrkflwPath):
        wrkbk = load_workbook(wrkflwPath)
        # Clean up current sheets in order to create new
        sheetAll = 'tblEncompassAllAct'
        sheetOpen = 'tblEncompassOpen'
        for sheet in wrkbk.sheetnames:
            if sheet == sheetAll:
                wrkbk.remove(wrkbk[sheet])
            elif sheet == sheetOpen:
                wrkbk.remove(wrkbk[sheet])
        # Create Excel Writer used to create tables
        writer = pd.ExcelWriter(wrkflwPath, mode='a',
                                datetime_format='MM-DD-YYYY', engine='openpyxl')
        writer.book = wrkbk
        # Create tblEncompassAllAct
        self.encmpDataAllAct.to_excel(writer, sheet_name='tblEncompassAllAct',
                                      startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassAllAct')
        table = Table(displayName='tblEncompassAllAct',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        # Create tblEncompassOpen
        self.encmpDataOpen.sort_values(by=['ApplicationDate'], inplace=True,
                                       ignore_index=True)
        self.encmpDataOpen['index'] = self.encmpDataOpen.index
        self.encmpDataOpen.to_excel(writer, sheet_name='tblEncompassOpen',
                                    startcol=0, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassOpen')
        table = Table(displayName='tblEncompassOpen',
                      ref='A1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        wrkbk.save(wrkflwPath)
        wrkbk.close()

    def dashData(self, wrkflwPath):
        # Open Workbook up and allow functions to compile
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workbook = excel.Workbooks.Open(wrkflwPath)
        workbook.Save()
        workbook.Close()
        excel.Quit()
        wrkbk = load_workbook(wrkflwPath)
        # Clean up current sheets in order to create new
        sheetDash = 'tblEncompassAllDash'
        for sheet in wrkbk.sheetnames:
            if sheet == sheetDash:
                wrkbk.remove(wrkbk[sheet])
        # Create Excel Writer used to create tables
        writer = pd.ExcelWriter(wrkflwPath, mode='a',
                                datetime_format='MM-DD-YYYY', engine='openpyxl')
        writer.book = wrkbk
        self.encmpDataAllExp = pd.read_excel(wrkflwPath, engine='openpyxl',
                                             sheet_name='tblEncompassAllExp')
        indexNamesExp = self.encmpDataAllExp[(self.encmpDataAllExp[
                                               'LoanNumber'] == 0) |
                                          (self.encmpDataAllExp['LoanNumber']
                                           == " ")].index
        self.encmpDataAllExp.drop(indexNamesExp, inplace=True)
        self.encmpDataAllExp[
             ['ApplicationDate', 'Disclosures', 'Processing', 'submittal',
              'Approval', 'ConditionSubmission', 'ClearToClose',
              'OS-FinancingContingency', 'LockExpirationDate', 'EstClosingDate',
              'ClosedDate']] = self.encmpDataAllExp[
             ['ApplicationDate', 'Disclosures', 'Processing', 'submittal',
              'Approval', 'ConditionSubmission', 'ClearToClose',
              'OS-FinancingContingency', 'LockExpirationDate', 'EstClosingDate',
              'ClosedDate']].apply(pd.to_datetime, errors='coerce')
        self.encmpDataAllExp['LoanStatus'] = np.where(
            self.encmpDataAllExp.ClosedDate.isnull(), 'Open', 'Closed')
        # Create tblEncompassAllDash
        self.encmpDataAllDash = pd.concat(
            [self.encmpDataAllAct, self.encmpDataAllExp])
        self.encmpDataAllDash = self.encmpDataAllDash.melt(
            id_vars=['Company-UsersOrganizationCode', 'LoanOfficer',
                     'LoanProcessor', 'BorrowerLastName', 'LoanNumber',
                     'LoanPurpose', 'LockRequestLoanAmount',
                     'LoanTeamMemberName-UW1-Initial', 'LoanStatus', 'DateType'
                     ], var_name='MilestoneType',
                        value_name='MilestoneDates')
        self.encmpDataAllDash['MilestoneOrder'] = [1 if x == 'Disclosures'
            else 2 if x == 'Processing' else 3 if x == 'submittal'
            else 4 if x == 'Approval' else 5 if x == 'ConditionSubmission'
            else 6 if x == 'ClearToClose'
            else 99 for x in self.encmpDataAllDash['MilestoneType']]
        self.encmpDataAllDash['MilestoneDates'].replace('', np.nan,
                                                        inplace=True)
        self.encmpDataAllDash.dropna(subset=['LoanNumber'], inplace=True)
        self.encmpDataAllDash['Lookup'] = self.encmpDataAllDash[
                                              'LoanNumber'].astype('int64').astype(str)+ \
                                          self.encmpDataAllDash['DateType'] + \
                                          self.encmpDataAllDash['MilestoneType']
        self.encmpDataAllDash.to_excel(writer, sheet_name='tblEncompassAllDash',
                                       startcol=1, index=False)
        sheet = wrkbk.get_sheet_by_name('tblEncompassAllDash')
        table = Table(displayName='tblEncompassAllDash',
                      ref='B1:' + get_column_letter(sheet.max_column) + str(
                          sheet.max_row))
        style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=False,
                               showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        wrkbk.save(wrkflwPath)
        wrkbk.close()
